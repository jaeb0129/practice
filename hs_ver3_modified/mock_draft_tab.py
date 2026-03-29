"""
KBO 모의 드래프트 — 독립 Streamlit 앱
  ·  10개 팀 × 11라운드 (기본값, 변경 가능)
  ·  선수 풀 드래그 → 보드 드롭
  ·  팀 순서 헤더 드래그로 변경
  ·  Excel 파일 업로드로 선수 명단 가져오기
  ·  이름/학교/포지션/투타 직접 추가
  ·  결과 Excel 내보내기 (2개 시트)

실행: streamlit run mock_draft_tab.py
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import io
from datetime import datetime

# ══════════════════════════════════════════════════════════════
# 상수
# ══════════════════════════════════════════════════════════════
KBO_TEAMS = [
    {"id": "WO", "name": "키움 히어로즈", "color": "#8B0029", "short": "키움"},
    {"id": "OB", "name": "두산 베어스",   "color": "#BBBBBB", "short": "두산"},
    {"id": "KI", "name": "KIA 타이거즈", "color": "#EA0029", "short": "KIA"},
    {"id": "LT", "name": "롯데 자이언츠", "color": "#002D9C", "short": "롯데"},
    {"id": "KT", "name": "KT 위즈",       "color": "#000000", "short": "KT"},
    {"id": "NC", "name": "NC 다이노스",   "color": "#1C5FAD", "short": "NC"},
    {"id": "SS", "name": "삼성 라이온즈", "color": "#074CA1", "short": "삼성"},
    {"id": "SK", "name": "SSG 랜더스",    "color": "#CE0E2D", "short": "SSG"},
    {"id": "HH", "name": "한화 이글스",   "color": "#FF6B00", "short": "한화"},
    {"id": "LG", "name": "LG 트윈스",     "color": "#C30452", "short": "LG"}
]

POSITIONS      = ["투수", "포수", "1루수", "2루수", "3루수", "유격수", "외야수", '우투', '좌투', '사이드']
HAND_OPTS      = ["우투우타", "우투좌타", "좌투좌타", "좌투우타", "우투양타"]
DEFAULT_ROUNDS = 11


# ══════════════════════════════════════════════════════════════
# 세션 초기화
# ══════════════════════════════════════════════════════════════
def _init():
    defaults = {
        "teams":    [t.copy() for t in KBO_TEAMS],
        "players":  [],
        "board":    {},
        "rounds":   DEFAULT_ROUNDS,
        "file_key": 0,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ══════════════════════════════════════════════════════════════
# Excel 파싱
# ══════════════════════════════════════════════════════════════
def _load_excel(f) -> list:
    """
    필수 컬럼: 이름 / name / 선수명
    선택 컬럼: 학교 / 소속 / school | 포지션 / position / pos | 투타 / hand
    """
    try:
        df = pd.read_excel(f, dtype=str)
        df.fillna("", inplace=True)

        col_map = {}
        for col in df.columns:
            lc = col.lower().strip()
            if "name" not in col_map   and any(k in lc for k in ["이름", "name", "선수"]):
                col_map["name"]     = col
            elif "school" not in col_map and any(k in lc for k in ["학교", "소속", "팀", "school", "org"]):
                col_map["school"]   = col
            elif "pos" not in col_map   and any(k in lc for k in ["포지션", "position", "pos"]):
                col_map["pos"]      = col
            elif "hand" not in col_map  and any(k in lc for k in ["투타", "hand", "throw", "bat"]):
                col_map["hand"]     = col

        if "name" not in col_map:
            st.error("❌ '이름' 컬럼 없음 — 컬럼명에 이름·name·선수 중 하나가 포함되어야 합니다.")
            return []

        out, seen = [], set()
        for i, row in df.iterrows():
            name = str(row[col_map["name"]]).strip()
            if not name or name.lower() == "nan" or name in seen:
                continue
            seen.add(name)
            out.append({
                "id":       f"xl_{i}_{name}",
                "name":     name,
                "school":   str(row.get(col_map.get("school", ""), "")).strip(),
                "position": str(row.get(col_map.get("pos",    ""), "")).strip(),
                "hand":     str(row.get(col_map.get("hand",   ""), "")).strip(),
                "drafted":  False,
            })
        return out
    except Exception as exc:
        st.error(f"❌ 파일 로드 실패: {exc}")
        return []


# ══════════════════════════════════════════════════════════════
# 드래프트 보드 HTML
# ══════════════════════════════════════════════════════════════
def _board_html(teams: list, players: list, rounds: int, board: dict) -> str:
    tj = json.dumps(teams,   ensure_ascii=False)
    pj = json.dumps(players, ensure_ascii=False)
    bj = json.dumps(board,   ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0;font-family:'Noto Sans KR',sans-serif}}
body{{background:#0d1117;color:#e6edf3;padding:10px 12px;font-size:13px}}

/* 상단 */
.topbar{{display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap}}
.badge{{background:#161b22;border:1px solid #30363d;padding:4px 12px;border-radius:20px;font-size:11px;color:#8b949e}}
.badge b{{color:#e6edf3}}
.btn{{padding:5px 13px;border-radius:6px;border:1px solid #30363d;background:#21262d;color:#e6edf3;cursor:pointer;font-size:12px;font-family:inherit;transition:all .15s}}
.btn:hover{{background:#30363d}}
.btn.save{{background:#1f6feb;border-color:#388bfd;font-weight:600}}
.btn.save:hover{{background:#388bfd}}
.btn.rst:hover{{background:#3d1a1a;border-color:#f85149;color:#f85149}}

/* 레이아웃 */
.layout{{display:grid;grid-template-columns:205px 1fr;gap:10px;align-items:start}}

/* 선수 풀 */
.pool{{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:10px;max-height:820px;overflow-y:auto;position:sticky;top:0}}
.pool-hd{{font-size:11px;color:#8b949e;text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid #30363d;display:flex;justify-content:space-between;align-items:center}}
.pcnt{{background:#21262d;border-radius:10px;padding:1px 7px;font-size:10px;color:#58a6ff}}
.search{{width:100%;background:#0d1117;border:1px solid #30363d;color:#e6edf3;padding:5px 8px;border-radius:5px;font-size:12px;margin-bottom:8px;outline:none}}
.search:focus{{border-color:#388bfd}}
.search::placeholder{{color:#484f58}}
.pcard{{background:#21262d;border:1px solid #30363d;border-radius:5px;padding:6px 8px;margin-bottom:4px;cursor:grab;user-select:none;transition:all .15s}}
.pcard:active{{cursor:grabbing;opacity:.6}}
.pcard:hover{{border-color:#58a6ff;background:#1c2333}}
.pcard.drafted{{opacity:.22;cursor:not-allowed;pointer-events:none}}
.pname{{font-weight:700;font-size:13px;line-height:1.3}}
.pmeta{{color:#8b949e;font-size:10px;margin-top:2px}}
.pbadge{{display:inline-block;background:#0d2457;color:#79c0ff;border-radius:3px;padding:1px 5px;font-size:9px;margin-right:3px;font-weight:600}}
.no-p{{color:#484f58;font-size:12px;text-align:center;padding:24px 0}}

/* 보드 */
.bwrap{{overflow-x:auto}}
table{{border-collapse:collapse;width:max-content;min-width:100%}}
thead th{{background:#161b22;border:1px solid #21262d;padding:0;position:sticky;top:0;z-index:20}}
thead th.rc{{width:44px;min-width:44px}}
thead th.tc{{width:118px;min-width:118px}}
.th-inner{{padding:9px 6px;text-align:center;cursor:grab;user-select:none;transition:background .15s}}
.th-inner:active{{cursor:grabbing}}
thead th:hover .th-inner{{background:#1c2333}}
thead th.drag-over{{outline:2px dashed #388bfd;outline-offset:-2px}}
.tname{{font-size:12px;font-weight:700;line-height:1.3}}
.thint{{font-size:9px;color:rgba(255,255,255,.28);margin-top:2px}}
.rhd{{text-align:center;color:#8b949e;font-size:11px;font-weight:700;vertical-align:middle}}

tbody tr:nth-child(odd) td{{background:#0d1117}}
tbody tr:nth-child(even) td{{background:#0f1318}}
tbody td{{border:1px solid #191e26;vertical-align:top;padding:3px}}
tbody td.rl{{background:#161b22!important;text-align:center;color:#8b949e;font-size:11px;font-weight:700;vertical-align:middle;position:sticky;left:0;z-index:5;border-right:2px solid #30363d;width:44px}}

/* 드롭존 */
.dz{{min-height:62px;border-radius:5px;border:1px dashed #21262d;display:flex;align-items:center;justify-content:center;transition:all .15s;padding:2px;color:#3d444d;font-size:10px}}
.dz.over{{border-color:#388bfd!important;background:rgba(56,139,253,.08);border-style:solid}}
.dz.filled{{border-color:transparent;padding:2px}}

/* 지명 카드 */
.dp{{background:#1c2333;border:1px solid #30363d;border-radius:5px;padding:5px 20px 5px 7px;width:100%;position:relative;transition:border-color .15s}}
.dp:hover{{border-color:#f85149}}
.dp-n{{font-weight:700;font-size:12px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.dp-m{{color:#8b949e;font-size:10px;margin-top:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.rmb{{position:absolute;top:4px;right:4px;background:none;border:none;color:#484f58;font-size:13px;cursor:pointer;line-height:1;padding:0;font-family:inherit}}
.rmb:hover{{color:#f85149}}

::-webkit-scrollbar{{width:4px;height:4px}}
::-webkit-scrollbar-track{{background:#0d1117}}
::-webkit-scrollbar-thumb{{background:#30363d;border-radius:2px}}
</style>
</head>
<body>

<div class="topbar">
  <span class="badge">⚾ KBO 모의 드래프트</span>
  <span class="badge" id="pickBadge"><b>0</b> / <b>0</b> 지명 완료</span>
  <button class="btn save" onclick="sendSave()">💾 저장</button>
  <button class="btn rst"  onclick="doReset()">🔄 보드 초기화</button>
</div>

<div class="layout">
  <div class="pool">
    <div class="pool-hd">선수 풀 <span class="pcnt" id="pcnt">0</span></div>
    <input class="search" type="text" placeholder="이름 · 학교 검색…" oninput="filterPool(this.value)">
    <div id="poolList"></div>
  </div>

  <div class="bwrap">
    <table id="dt">
      <thead><tr id="hrow"><th class="rc rhd">R</th></tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</div>

<script>
let teams     = {tj};
let players   = {pj};
let board     = {bj};
let rounds    = {rounds};
let teamOrder = teams.map(t => t.id);

let dragPid = null;
let dragTid = null;

// ── 렌더 ──
function render() {{
  renderHeader();
  renderRows();
  renderPool();
  updateBadge();
}}

function renderHeader() {{
  const row = document.getElementById("hrow");
  while (row.children.length > 1) row.removeChild(row.lastChild);
  teamOrder.forEach(tid => {{
    const team = teams.find(t => t.id === tid);
    const th = document.createElement("th");
    th.className = "tc";
    th.dataset.tid = tid;
    const accent = (team.color === "#FFFFFF" || team.color === "#BBBBBB") ? "#888888" : team.color;
    th.innerHTML = `<div class="th-inner" style="border-top:3px solid ${{accent}}">
      <div class="tname">${{team.short}}</div>
      <div class="thint">⠿ 순서 변경</div>
    </div>`;
    th.addEventListener("dragstart",  thDS);
    th.addEventListener("dragover",   thDO);
    th.addEventListener("dragleave",  thDL);
    th.addEventListener("drop",       thDrop);
    th.draggable = true;
    row.appendChild(th);
  }});
}}

function renderRows() {{
  const tbody = document.getElementById("tbody");
  tbody.innerHTML = "";
  for (let r = 1; r <= rounds; r++) {{
    const tr = document.createElement("tr");
    const tdL = document.createElement("td");
    tdL.className = "rl";
    tdL.textContent = r + "R";
    tr.appendChild(tdL);

    teamOrder.forEach(tid => {{
      const pick = getPick(r, tid);
      const td   = document.createElement("td");
      const dz   = document.createElement("div");
      dz.className = "dz" + (pick ? " filled" : "");
      dz.dataset.round = r;
      dz.dataset.tid   = tid;

      if (pick) {{
        dz.innerHTML = `<div class="dp">
          <button class="rmb" onclick="removePick(${{r}},'${{tid}}')">✕</button>
          <div class="dp-n">${{pick.name}}</div>
          <div class="dp-m">${{pick.position ? pick.position + " · " : ""}}${{pick.school || ""}}</div>
        </div>`;
      }} else {{
        dz.textContent = "드롭";
        dz.addEventListener("dragover",  dzO);
        dz.addEventListener("dragleave", dzL);
        dz.addEventListener("drop",      dzDrop);
      }}
      td.appendChild(dz);
      tr.appendChild(td);
    }});
    tbody.appendChild(tr);
  }}
}}

function renderPool(filter) {{
  filter = filter || "";
  const list = document.getElementById("poolList");
  list.innerHTML = "";
  const lc = filter.toLowerCase();
  const vis = players.filter(p =>
    !p.drafted &&
    (!filter || p.name.toLowerCase().includes(lc) || (p.school||"").toLowerCase().includes(lc))
  );
  document.getElementById("pcnt").textContent = vis.length;
  if (!vis.length) {{
    list.innerHTML = '<div class="no-p">선수 없음</div>';
    return;
  }}
  vis.forEach(p => {{
    const d = document.createElement("div");
    d.className = "pcard";
    d.draggable = true;
    d.dataset.pid = p.id;
    d.innerHTML = `<div class="pname">${{p.name}}</div>
      <div class="pmeta">
        ${{p.position ? `<span class="pbadge">${{p.position}}</span>` : ""}}
        ${{p.school || ""}}${{p.hand ? " · " + p.hand : ""}}
      </div>`;
    d.addEventListener("dragstart", pdDS);
    list.appendChild(d);
  }});
}}

function updateBadge() {{
  let done = 0, total = rounds * teamOrder.length;
  for (let r = 1; r <= rounds; r++)
    teamOrder.forEach(tid => {{ if (getPick(r, tid)) done++; }});
  document.getElementById("pickBadge").innerHTML =
    `<b>${{done}}</b> / <b>${{total}}</b> 지명 완료`;
}}

// ── 헬퍼 ──
function getPick(r, tid) {{
  return (board[r] && board[r][tid]) || (board[String(r)] && board[String(r)][tid]) || null;
}}

// ── 선수 D&D ──
function pdDS(e) {{ dragPid = e.currentTarget.dataset.pid; e.dataTransfer.effectAllowed = "move"; }}
function dzO(e)  {{ e.preventDefault(); e.currentTarget.classList.add("over"); }}
function dzL(e)  {{ e.currentTarget.classList.remove("over"); }}
function dzDrop(e) {{
  e.preventDefault();
  const dz = e.currentTarget;
  dz.classList.remove("over");
  const r   = parseInt(dz.dataset.round);
  const tid = dz.dataset.tid;
  const p   = players.find(x => x.id === dragPid);
  if (!p || p.drafted) return;
  if (!board[r]) board[r] = {{}};
  board[r][tid] = {{...p}};
  p.drafted = true;
  dragPid = null;
  sendUpdate();
  render();
}}
function removePick(r, tid) {{
  const pick = getPick(r, tid);
  if (!pick) return;
  if (board[r])        board[r][tid]        = null;
  if (board[String(r)]) board[String(r)][tid] = null;
  const p = players.find(x => x.id === pick.id);
  if (p) p.drafted = false;
  sendUpdate();
  render();
}}

// ── 팀 헤더 D&D ──
function thDS(e)  {{ dragTid = e.currentTarget.dataset.tid; e.dataTransfer.effectAllowed = "move"; }}
function thDO(e)  {{ e.preventDefault(); e.currentTarget.classList.add("drag-over"); }}
function thDL(e)  {{ e.currentTarget.classList.remove("drag-over"); }}
function thDrop(e) {{
  e.preventDefault();
  e.currentTarget.classList.remove("drag-over");
  const toTid = e.currentTarget.dataset.tid;
  if (!dragTid || dragTid === toTid) return;
  const fi = teamOrder.indexOf(dragTid);
  const ti = teamOrder.indexOf(toTid);
  teamOrder.splice(fi, 1);
  teamOrder.splice(ti, 0, dragTid);
  dragTid = null;
  sendUpdate();
  render();
}}

// ── 검색 / 초기화 ──
function filterPool(v) {{ renderPool(v); }}
function doReset() {{
  if (!confirm("보드를 초기화할까요?\\n선수 목록은 그대로 유지됩니다.")) return;
  board = {{}};
  players.forEach(p => p.drafted = false);
  sendUpdate();
  render();
}}

// ── Streamlit 통신: 클립보드 복사 방식 ──
// components.html()은 값 반환이 안 되므로
// 저장 버튼 클릭 시 JSON을 클립보드에 복사 → 유저가 text_area에 붙여넣기
function sendUpdate(action) {{
  // 내부 상태만 갱신 (UI 즉시 반영)
}}
function sendSave() {{
  const payload = JSON.stringify({{ board, teamOrder, players }});
  // 클립보드 복사
  if (navigator.clipboard && navigator.clipboard.writeText) {{
    navigator.clipboard.writeText(payload).then(() => {{
      showToast("📋 JSON 복사 완료! Streamlit 텍스트창에 붙여넣기(Ctrl+V) 후 [상태 저장] 클릭");
    }}).catch(() => {{ fallbackCopy(payload); }});
  }} else {{
    fallbackCopy(payload);
  }}
}}
function fallbackCopy(text) {{
  const ta = document.createElement("textarea");
  ta.value = text;
  ta.style.position = "fixed"; ta.style.opacity = "0";
  document.body.appendChild(ta);
  ta.focus(); ta.select();
  try {{ document.execCommand("copy"); showToast("📋 복사 완료! 붙여넣기(Ctrl+V) 후 [상태 저장] 클릭"); }}
  catch(e) {{ showToast("❌ 복사 실패 — 브라우저 권한을 확인하세요"); }}
  document.body.removeChild(ta);
}}
function showToast(msg) {{
  const t = document.createElement("div");
  t.textContent = msg;
  t.style.cssText = "position:fixed;bottom:20px;left:50%;transform:translateX(-50%);"
    + "background:#1f6feb;color:#fff;padding:10px 18px;border-radius:8px;"
    + "font-size:12px;z-index:9999;box-shadow:0 4px 12px rgba(0,0,0,.4);max-width:90%;text-align:center";
  document.body.appendChild(t);
  setTimeout(() => t.remove(), 4000);
}}

render();
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════
# Excel 내보내기
# ══════════════════════════════════════════════════════════════
def _export_excel(board: dict, teams: list, team_order: list, rounds: int) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tmap = {t["id"]: t for t in teams}

    def _pick(r, tid):
        return (
            (board.get(str(r)) or {}).get(tid)
            or (board.get(r) or {}).get(tid)
        )

    # 시트1: 전체 목록
    rows = []
    for r in range(1, rounds + 1):
        for idx, tid in enumerate(team_order):
            pick = _pick(r, tid)
            rows.append({
                "라운드":    r,
                "전체 순번": (r - 1) * len(team_order) + idx + 1,
                "팀":        tmap.get(tid, {}).get("name", tid),
                "선수 이름": pick["name"]             if pick else "",
                "포지션":    pick.get("position", "") if pick else "",
                "소속/학교": pick.get("school",   "") if pick else "",
                "투타":      pick.get("hand",     "") if pick else "",
            })
    df_list = pd.DataFrame(rows)

    # 시트2: 팀별 피벗
    pivot = {}
    for r in range(1, rounds + 1):
        rd = {}
        for tid in team_order:
            pick = _pick(r, tid)
            rd[tmap.get(tid, {}).get("name", tid)] = pick["name"] if pick else ""
        pivot[f"{r}라운드"] = rd
    df_pivot = pd.DataFrame(pivot).T
    df_pivot.index.name = "라운드"

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_list.to_excel(writer,  sheet_name="전체 지명 목록",    index=False)
        df_pivot.to_excel(writer, sheet_name="팀별 라운드 요약")

        wb = writer.book

        def thin():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        # ── 밝은 테마 스타일 ──
        H_FILL  = PatternFill("solid", fgColor="1F4E79")   # 짙은 파랑 헤더
        H_FONT  = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
        B_FONT  = Font(size=10, name="맑은 고딕", color="000000")
        CENTER  = Alignment(horizontal="center", vertical="center")
        ODD     = PatternFill("solid", fgColor="FFFFFF")   # 흰색
        EVEN    = PatternFill("solid", fgColor="DCE6F1")   # 연한 파랑

        for ws_name in ["전체 지명 목록", "팀별 라운드 요약"]:
            ws = wb[ws_name]
            for cell in ws[1]:
                cell.fill = H_FILL; cell.font = H_FONT
                cell.alignment = CENTER; cell.border = thin()
            for ri in range(2, ws.max_row + 1):
                fill = ODD if ri % 2 == 1 else EVEN
                for cell in ws[ri]:
                    cell.fill = fill; cell.font = B_FONT
                    cell.alignment = CENTER; cell.border = thin()
            for col_cells in ws.columns:
                w = max((len(str(c.value or "")) for c in col_cells), default=6)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w * 2.2, 30)
            ws.row_dimensions[1].height = 22
            for ri in range(2, ws.max_row + 1):
                ws.row_dimensions[ri].height = 18
            ws.freeze_panes = "B2"

        wb["전체 지명 목록"].sheet_properties.tabColor    = "1F6FEB"
        wb["팀별 라운드 요약"].sheet_properties.tabColor  = "3FB950"

    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# 메인 탭 함수
# ══════════════════════════════════════════════════════════════
def render_mock_draft_tab():
    _init()

    # 헤더
    st.markdown(
        "<h2 style='margin-bottom:4px'>⚾ KBO 모의 드래프트</h2>"
        "<p style='color:#8b949e;font-size:13px;margin-bottom:16px'>"
        "선수 카드를 드래그해서 팀·라운드 칸에 배치하세요. "
        "팀 헤더도 드래그로 순서 변경 가능합니다.</p>",
        unsafe_allow_html=True,
    )

    # ── 컨트롤 패널 ──
    c1, c2, c3, c4 = st.columns([1, 2, 2, 2])

    with c1:
        rounds = st.number_input(
            "라운드 수", min_value=1, max_value=20,
            value=st.session_state.rounds, step=1,
        )
        if rounds != st.session_state.rounds:
            st.session_state.rounds = rounds

    with c2:
        st.markdown("**📂 선수 명단 (Excel)**")
        uploaded = st.file_uploader(
            "xlsx", type=["xlsx", "xls"],
            label_visibility="collapsed",
            key=f"uploader_{st.session_state.file_key}",
        )
        if uploaded is not None:
            loaded = _load_excel(uploaded)
            if loaded:
                exist = {p["name"] for p in st.session_state.players}
                new   = [p for p in loaded if p["name"] not in exist]
                st.session_state.players.extend(new)
                if new:
                    st.success(f"✅ {len(new)}명 추가 (총 {len(st.session_state.players)}명)")
                else:
                    st.info("이미 모두 등록된 선수입니다.")

    with c3:
        st.markdown("**✏️ 선수 직접 추가**")
        with st.expander("➕ 폼 열기", expanded=False):
            r1a, r1b = st.columns(2)
            n_name   = r1a.text_input("이름 *",    key="man_name",   placeholder="홍길동")
            n_school = r1b.text_input("학교/소속", key="man_school", placeholder="○○고등학교")
            r2a, r2b = st.columns(2)
            n_pos    = r2a.selectbox("포지션", [""] + POSITIONS, key="man_pos")
            n_hand   = r2b.selectbox("투타",   [""] + HAND_OPTS,   key="man_hand")
            if st.button("추가", key="btn_add", type="primary"):
                nm = n_name.strip()
                if not nm:
                    st.warning("이름을 입력해 주세요.")
                elif nm in {p["name"] for p in st.session_state.players}:
                    st.warning(f"'{nm}'은(는) 이미 등록되어 있습니다.")
                else:
                    st.session_state.players.append({
                        "id":       f"man_{nm}_{len(st.session_state.players)}",
                        "name":     nm,
                        "school":   n_school.strip(),
                        "position": n_pos,
                        "hand":     n_hand,
                        "drafted":  False,
                    })
                    st.success(f"'{nm}' 추가 완료!")
                    st.rerun()

    with c4:
        st.markdown("**📥 내보내기**")

        team_order_ids = [t["id"] for t in st.session_state.teams]
        xl = _export_excel(
            board=st.session_state.board,
            teams=st.session_state.teams,
            team_order=team_order_ids,
            rounds=st.session_state.rounds,
        )
        st.download_button(
            label="📥 Excel 다운로드",
            data=xl,
            file_name=f"KBO_모의드래프트_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    st.divider()

    # 안내
    if not st.session_state.players:
        st.info(
            "📂 **Excel 파일을 업로드**하거나 직접 추가해 주세요.\n\n"
            "| 구분 | 인식 키워드 |\n|---|---|\n"
            "| 필수 | `이름` `name` `선수` |\n"
            "| 선택 | `학교` `소속` `포지션` `투타` |"
        )

    # ── 상태 브릿지 안내 ──
    # components.html()은 값을 반환하지 않으므로
    # 보드의 「💾 저장」 버튼 클릭 시 JSON이 클립보드에 복사됨
    # → 아래 텍스트창에 붙여넣기 → 「상태 저장」 클릭
    st.markdown(
        "<p style='font-size:11px;color:#8b949e;margin-top:4px;margin-bottom:4px'>"
        "💡 드래그 완료 후 보드의 <b>「💾 저장」</b> 버튼 클릭 → 아래 창에 JSON이 복사됩니다 "
        "→ 창에 붙여넣기(Ctrl+V) → <b>「상태 저장」</b> 클릭하면 Excel에 반영됩니다.</p>",
        unsafe_allow_html=True,
    )
    b_col1, b_col2 = st.columns([6, 1])
    with b_col1:
        raw_json = st.text_area(
            "draft_bridge",
            value="",
            height=68,
            label_visibility="collapsed",
            placeholder="보드에서 「💾 저장」 클릭 후 여기에 붙여넣기(Ctrl+V) → 「상태 저장」 클릭",
            key="draft_json_input",
        )
    with b_col2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        save_clicked = st.button("상태 저장 ✅", use_container_width=True, type="primary", key="btn_state_save")

    if save_clicked and raw_json and raw_json.strip():
        try:
            data = json.loads(raw_json.strip())
            if isinstance(data, dict):
                if data.get("board") is not None:
                    st.session_state.board = data["board"]
                new_order = data.get("teamOrder", [])
                if new_order:
                    id_map  = {t["id"]: t for t in st.session_state.teams}
                    ordered = [id_map[tid] for tid in new_order if tid in id_map]
                    if ordered:
                        st.session_state.teams = ordered
                pid_map = {p["id"]: p for p in st.session_state.players}
                for pd_item in data.get("players", []):
                    pid = pd_item.get("id")
                    if pid in pid_map:
                        pid_map[pid]["drafted"] = bool(pd_item.get("drafted"))
                st.success("✅ 보드 상태 저장 완료! Excel 다운로드를 클릭하세요.")
                st.rerun()
        except json.JSONDecodeError:
            st.error("❌ JSON 형식 오류입니다. 보드에서 「💾 저장」을 다시 클릭해주세요.")

    # ── 드래프트 보드 ──
    h = max(52 + st.session_state.rounds * 68 + 80, 480)
    html = _board_html(
        teams=st.session_state.teams,
        players=st.session_state.players,
        rounds=st.session_state.rounds,
        board=st.session_state.board,
    )
    components.html(html, height=h, scrolling=True)

    # ── 하단 요약 테이블 ──
    if st.session_state.board:
        has_picks = any(
            bool(v) for rv in st.session_state.board.values()
            for v in (rv.values() if isinstance(rv, dict) else [])
        )
        if has_picks:
            st.markdown("#### 📋 지명 현황")
            tmap      = {t["id"]: t["name"] for t in st.session_state.teams}
            order_ids = [t["id"] for t in st.session_state.teams]
            pivot = {}
            for r in range(1, st.session_state.rounds + 1):
                rd = {}
                for tid in order_ids:
                    pick = (
                        (st.session_state.board.get(str(r)) or {}).get(tid)
                        or (st.session_state.board.get(r) or {}).get(tid)
                    )
                    rd[tmap.get(tid, tid)] = pick["name"] if pick else "—"
                pivot[f"{r}R"] = rd
            pv_df = pd.DataFrame(pivot).T
            pv_df.index.name = "라운드"
            st.dataframe(pv_df, use_container_width=True,
                         height=min(420, 35 * (st.session_state.rounds + 2)))


# ══════════════════════════════════════════════════════════════
# 독립 실행
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    st.set_page_config(
        page_title="KBO 모의 드래프트",
        page_icon="⚾",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown("""
    <style>
    .stApp { background-color: #0d1117; color: #e6edf3; }
    section[data-testid="stSidebar"] { display: none; }
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    </style>
    """, unsafe_allow_html=True)

    render_mock_draft_tab()
